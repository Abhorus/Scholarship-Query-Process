import openpyxl 
import os
import datetime
from fuzzywuzzy import fuzz

start_time = datetime.datetime.now()
os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries\FY21")

wb = openpyxl.load_workbook('2021_03_03_OSF_SCHOLARSHIP_PSTD_ENROLMNT.xlsx')
#wb = openpyxl.load_workbook('_Altloan_Testfile_2020_11_04_OSF_SCHOLARSHIP_PSTD_ENROLMNT.xlsx')
sheet = wb.active
ScholarshipItemTypes = ['050000000014','050000000016','050000000019','050000000022']
unappliedwb = openpyxl.load_workbook('2021_03_03_OSF_UNAPPLIED_CREDITS_FILTER.xlsx')
enrollmentHours = openpyxl.load_workbook('2021_03_03_OSF_SCHOLARSHIP_ENROLMNT_HOURS.xlsx')
enrollmentHourSheet = enrollmentHours.active
unappliedSheet = unappliedwb.active
current_date = datetime.date.today()


####vlookup####
os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Exceptions")
exceptions = openpyxl.load_workbook('1212 Exceptions.xlsx')
execsheet = exceptions.active
vlookup = set()
for i in list(execsheet.columns)[0]:
    #vlookup.append(str(i.value))
    vlookup.add(str(i.value))
    #print(i.value, type(i.value))
    
#for i, e in enumerate(vlookup):
    #print(e)

#print(vlookup, len(vlookup))
###creating a new spreadsheet to save results
resultswb = openpyxl.Workbook()
resultsSheet = resultswb.active #sheet 1
resultsSheet.title = 'PSTD Results'
resultswb.create_sheet(index=1, title= 'Unapplied Results')
resultswb.create_sheet(index=2, title= 'Enrollment Hour Results')
resultsSheet2 = resultswb['Unapplied Results']
resultsSheet3 = resultswb['Enrollment Hour Results']

#resultsSheet2.cell(row=1, column=1).value = 'ID'
#resultsSheet2.cell(row=1, column=2).value = 'Item Type'
#resultsSheet2.cell(row=1, column=3).value = 'Descr'
#resultsSheet2.cell(row=1, column=4).value = 'Item Amt'
#resultsSheet2.cell(row=1, column=5).value = 'Term'
#resultsSheet2.cell(row=1, column=6).value = 'Take Prgrs'
#resultsSheet2.cell(row=1, column=7).value = 'Career'
#resultsSheet2.cell(row=1, column=8).value = 'Ref Nbr'
#resultsSheet2.cell(row=1, column=9).value = 'Postd DtTm'
#resultsSheet2.cell(row=1, column=10).value = 'User'


headerList = ['ID','Item Type','Descr','Item Amt','Term','Take Prgrs','Career','Ref Nbr','Postd DtTm','User']
for i, ele in enumerate(headerList):
    resultsSheet.cell(row=1, column= i+1).value = ele
    resultsSheet2.cell(row=1, column= i+1).value = ele
    resultsSheet3.cell(row=1, column= i+1).value = ele

resultsSheet2.cell(row=1, column=11).value = 'Take Prgrs' #orginal spreadsheet has different headers


###PSTD Query####
print('from pstd query')
count = 0

a = list(sheet.columns)[5] #list of column 'Take Prgrs'/credit hours of each student
for i in range(len(a)): #loops through each element and checks to see if it is equal to 0 
    if a[i].value == 0: #problem here with iknowican as it will pull in the students < fulltime that are returned to donor;; add comparsion to exceptions list?
        #print(sheet.cell(row=i+1, column=1).value)
        if sheet.cell(row=i+1, column=1).value not in vlookup: 
            count += 1
            #print(sheet.cell(row=i+1, column=1).value)
            for index, ele in enumerate(list(sheet.rows)[i]): #if element in 'a' is 0, grabs every element in row i
                print(ele.value, end=" ")
                resultsSheet.cell(row= count+1, column=index+1).value = ele.value
            print('\n')
            
###IKIC Logic###
ikicList = [
'I KNOW I CAN',
'IKIC',
]
#needs to check exceptions list
for listitem in ikicList:
    for i, ele in enumerate(list(sheet.columns)[7]):
        if fuzz.token_set_ratio(listitem, ele.value) > 90 and int(sheet.cell(row=i+1, column=6).value) < 12 and sheet.cell(row=i+1, column=1).value not in vlookup:
           count += 1
           for j, elej in enumerate(list(sheet.rows)[i]):
                print(elej.value, end=" ")
                resultsSheet.cell(row= count+1, column=j+1).value = elej.value
           print('\n')
  
    
print('Count = ', count, datetime.date.today())
if count == 0:
    resultsSheet.cell(row= count+2, column=1).value = 'No Results'


####unapplied query####
unappliedCount = 0
for index, i in enumerate(list(unappliedSheet.columns)[3]):
    if i.value in ScholarshipItemTypes and unappliedSheet.cell(row=index+1, column=11).value == 0 :
        #print(i.value, index, unappliedSheet.cell(row=index+1, column=1).value)
        unappliedCount += 1
        for indexj, ele in enumerate(list(unappliedSheet.rows)[index]):
            print(ele.value, end=" ")  
            resultsSheet2.cell(row = unappliedCount+1, column =indexj+1).value = ele.value
            #adds values to 1st row and each adjacent column in new spreadsheet
        print('\n')
       

if unappliedCount == 0:
    resultsSheet2.cell(row = unappliedCount+2, column =1).value = 'No Results'
"""need to compare to exceptions list """

####enrollment hour query####
hItemTypes = ['050000000033',
'050000000035',
'050000000021',
'050000000023'
]
enrollmentCount = 0
for i, ele in enumerate(list(enrollmentHourSheet.columns)[5]):
    #print(ele.value)
    if ele.value == None or type(ele.value) == str:
        continue
    elif enrollmentHourSheet.cell(row=i+1, column= 2).value in hItemTypes and ele.value > 0:
        enrollmentCount += 1
        for j, elej in enumerate(list(enrollmentHourSheet.rows)[i]):
            print(elej.value, end=" ")
            resultsSheet3.cell(row= enrollmentCount+1, column=j+1).value = elej.value
        print('\n')

if enrollmentCount == 0:
    resultsSheet3.cell(row= enrollmentCount+2, column = 1).value = "No Results"
    
os.chdir(r"K:\BF\OFS\Bursar\Processing\_External Payments\Scholarships\Queries\Enrollment Queries\Query Results")
resultswb.save('Query Results_' + str(current_date) + '.xlsx')  
print('Elapsed: ',datetime.datetime.now() - start_time )


  
###loops through all values in spreadsheet   
#for i in sheet:
#    for j in i:
#       print(j.value)