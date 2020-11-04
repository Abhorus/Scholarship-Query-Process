import openpyxl 
import os
import datetime

os.chdir(r"") #target location of where the spreadsheet is saved

wb = openpyxl.load_workbook('2020_11_04_OSF_SCHOLARSHIP_PSTD_ENROLMNT.xlsx') #example of spreadsheet name
#sheet = wb['Sheet1']
sheet = wb.active
ScholarshipItemTypes = ['050000000014','050000000016','050000000019','050000000022']
unappliedwb = openpyxl.load_workbook('2020_11_04_OSF_UNAPPLIED_CREDITS_FILTER.xlsx') #example of spreadsheet name
unappliedSheet = unappliedwb.active
current_date = datetime.date.today()


###vlookup

os.chdir(r"") #target location of spreadsheet that contains the data needed to do the comparision.
exceptions = openpyxl.load_workbook('1208 Exceptions.xlsx')
execsheet = exceptions.active
vlookup = set()
for i in list(execsheet.columns)[0]:
    vlookup.add(str(i.value))


###creating a new spreadsheet to save results

resultswb = openpyxl.Workbook()
resultsSheet = resultswb.active #sheet 1
resultsSheet.title = 'PSTD Results'
resultswb.create_sheet(index=1, title= 'Unapplied Results')
resultsSheet2 = resultswb['Unapplied Results']


###PSTD Query####
print('from pstd query')
count = 0

a = list(sheet.columns)[5] #list of column 'Take Prgrs'/credit hours of each student
for i in range(len(a)): #loops through each element and checks to see if it is equal to 0 
    if a[i].value == 0: #could add IKIC logic here: 'or a[i] < 12 and row i, ref column == i know i can, ikic

        if str(sheet.cell(row=i+1, column=1).value) not in vlookup:
            
            count += 1
            #print(sheet.cell(row=i+1, column=1).value)
            for index, ele in enumerate(list(sheet.rows)[i]): #if element in 'a' is 0, grabs every element in row i
                print(ele.value, end=" ")
                resultsSheet.cell(row= count, column=index+1).value = ele.value

            print('\n')
           

print('Count = ', count, datetime.date.today())
              
        
####unapplied query####
unappliedCount = 0
for index, i in enumerate(list(unappliedSheet.columns)[3]):
        if i.value in ScholarshipItemTypes and unappliedSheet.cell(row=index+1, column=11).value ==0 :
            #print(i.value, index, unappliedSheet.cell(row=index+1, column=1).value)
            unappliedCount += 1
            for indexj, ele in enumerate(list(unappliedSheet.rows)[index]):
                print(ele.value, indexj, end=" ")
                
                resultsSheet2.cell(row = unappliedCount, column =indexj+1).value = ele.value
                #adds values to 1st row and each adjacent column in new spreadsheet
            print('\n')
            

    
    
os.chdir(r"drive/Scholarships\Queries\Enrollment Queries\Query Results") #example of where to save the query results
resultswb.save('Query Results_' + str(current_date) + '.xlsx')    
###loops through all values in spreadsheet   
#for i in sheet:
#    for j in i:
#       print(j.value)